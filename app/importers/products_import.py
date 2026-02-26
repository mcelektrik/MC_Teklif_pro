from __future__ import annotations
import hashlib
import re

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook, Workbook

from app.importers.excel_utils import norm_str, norm_currency, norm_float, is_note_row
from app.core.db import SessionLocal
from app.core.schema import Product

REQUIRED = ["stok_kodu", "ad", "birim", "kdv_oran", "satis_fiyat", "para_birimi"]

@dataclass
class RowError:
    row_no: int
    stok_kodu: str
    reason: str

def _make_error_xlsx(errors: list[RowError], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "ERRORS"
    ws.append(["row_no", "stok_kodu", "reason"])
    for e in errors:
        ws.append([e.row_no, e.stok_kodu, e.reason])
    wb.save(out_path)

def import_products_from_excel(xlsx_path: str, sheet_name: str | None = None) -> dict[str, Any]:
    path = Path(xlsx_path).resolve()
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    def clean_hdr(h: str | None) -> str | None:
        h = norm_str(h)
        if not h:
            return None
        # drop anything after '*' and remove parenthetical hints: 'birim* (..)' -> 'birim'
        h = h.split('*', 1)[0].strip()
        h = re.sub(r'\s*\(.*?\)\s*', '', h).strip()
        return h or None
    headers = [clean_hdr(c.value) for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers) if h}

    errors: list[RowError] = []
    ok = skipped = inserted = updated = 0

    def get(row, key):
        i = idx.get(key)
        return row[i] if i is not None and i < len(row) else None

    db = SessionLocal()
    try:
        for excel_row, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(r):
                continue

            stok_kodu = norm_str(get(r, "stok_kodu"))

            # SKIP_TEMPLATE_NOTES: ignore instruction rows in template (NOTLAR/Kurallar...)
            if (not stok_kodu) or (stok_kodu.strip().upper() == "NOTLAR") or stok_kodu.strip().lower().startswith(("zorunlu", "kurallar", "para_birimi", "kdv_oran")):
                skipped += 1
                continue

            if is_note_row(stok_kodu):
                skipped += 1
                continue

            missing = []
            for k in REQUIRED:
                if k not in idx:
                    missing.append(f"kolon yok: {k}")
                else:
                    if norm_str(get(r, k)) == "":
                        missing.append(f"alan bos: {k}")
            if missing:
                errors.append(RowError(excel_row, stok_kodu, "; ".join(missing)))
                continue

            name = norm_str(get(r, "ad"))
            unit = norm_str(get(r, "birim"))
            kdv = norm_float(get(r, "kdv_oran"))
            sale = norm_float(get(r, "satis_fiyat"))
            cur = norm_currency(get(r, "para_birimi"))

            if kdv is None:
                errors.append(RowError(excel_row, stok_kodu, "kdv_oran sayi degil"))
                continue
            if sale is None:
                errors.append(RowError(excel_row, stok_kodu, "satis_fiyat sayi degil"))
                continue
            if cur not in ("TRY", "USD", "EUR"):
                errors.append(RowError(excel_row, stok_kodu, f"para_birimi gecersiz: {cur}"))
                continue

            # upsert - DB'de farkliysa yaz, degilse dokunma
            p = db.query(Product).filter(Product.stock_code == stok_kodu).one_or_none()
            if p:
                changed = False
                if p.name != name: p.name = name; changed = True
                if p.unit != unit: p.unit = unit; changed = True
                if p.vat_rate != kdv: p.vat_rate = kdv; changed = True
                if p.currency != cur: p.currency = cur; changed = True
                if p.sale_price != sale: p.sale_price = sale; changed = True
                if changed: updated += 1
            else:
                p = Product(
                    stock_code=stok_kodu,
                    name=name,
                    unit=unit,
                    vat_rate=kdv,
                    currency=cur,
                    sale_price=sale,
                )
                db.add(p)
                inserted += 1

            ok += 1

        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()

    out_err = None
    if errors:
        out_err = Path.cwd() / f"_import_errors_products_{path.stem}.xlsx"
        _make_error_xlsx(errors, out_err)

    return {
        "file": str(path),
        "file_sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
        "ok_rows": ok,
        "skipped_rows": skipped,
        "inserted": inserted,
        "updated": updated,
        "error_rows": len(errors),
        "error_report": str(out_err) if out_err else None,
    }

if __name__ == "__main__":
    import sys, json
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python app/importers/products_import.py <xlsx_path>")
    res = import_products_from_excel(sys.argv[1])
    print(json.dumps(res, ensure_ascii=False, indent=2))
